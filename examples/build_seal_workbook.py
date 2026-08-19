# -------------------------------------*- vatic -*----------------------------
#                               Open Source Risk Analysis
#
#                             Copyright (c) 2026, eggzec
#                          Contact: https://eggzec.github.io/
#
#                         License: GNU General Public License
#                              Version 3, 29 June 2007
#
# ----------------------------------------------------------------------------
#
#  Author(s)
#      Saud Zahir <m.saud.zahir@gmail.com>
#
#  Date
#      7 May 2026
#
#  Description
#      Generate the Double-D seal gland workbook used by the spreadsheet
#      examples and tests.
#
# ----------------------------------------------------------------------------
#
#  The original vatic project by Abraham Lee
#  (https://github.com/tisimst/vatic) ships an equivalent workbook, but that
#  repository carries no licence, so its file is not redistributed here.
#  This script rebuilds an equivalent model from the same published
#  dimensions and formulas, which keeps the example self-contained and lets
#  the workbook be regenerated or edited from source control.
#
#  Requires openpyxl:
#
#      uv pip install openpyxl
#      python examples/build_seal_workbook.py
#
# ----------------------------------------------------------------------------

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName


OUTPUT = Path(__file__).resolve().parent / "seal_tolerance.xlsx"

SHEET = "Seal-Groove Design"

#: Row, label, nominal, tolerance, units and the workbook name bound to the
#: nominal cell so the formulas below stay readable.
INPUTS: tuple[tuple[int, str, float, float, str, str], ...] = (
    (8, "Seal Height (mean)", 0.162, 0.005, "in", "sh"),
    (9, "Seal Width", 0.118, 0.005, "in", "sw"),
    (10, "Core Hole Diameter", 0.071, 0.005, "in", "d"),
    (11, "Core Hole % Compression", 0.75, 0.10, "", "cc"),
    (12, "Groove Height", 0.117, 0.002, "in", "gh"),
    (13, "Groove Width", 0.130, 0.002, "in", "gw"),
    (14, "Lid Flatness (GD&T)", 0.0, 0.005, "in", "flat1"),
    (15, "Box Flatness (GD&T)", 0.0, 0.005, "in", "flat2"),
)

#: Row, label, formula, units and the name bound to the result cell.
INTERMEDIATES: tuple[tuple[int, str, str, str, str], ...] = (
    (24, "Core Hole Area", "=cc*PI()*(d/2)^2", "in^2", "ca"),
    (25, "Seal Area", "=PI()*(sw/2/2)^2+(sw*(sh-sw/4))-ca", "in^2", "sa"),
    (26, "Groove Area", "=gw*(gh+flat1+flat2)", "in^2", "ga"),
)

#: Row, label, formula and the spec limits the characteristic is judged on.
OUTPUTS: tuple[tuple[int, str, str, float, float], ...] = (
    (30, "Gland Fill %", "=sa/ga", 0.75, 1.00),
    (31, "Seal Comp. %", "=1-gh/sh", 0.25, 0.50),
)

_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_HEADING = Font(bold=True, color="2323FF")
_BOLD = Font(bold=True)


def build() -> Path:
    """Write the workbook to :data:`OUTPUT`.

    Returns:
        The path the workbook was written to.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = SHEET

    sheet["B2"] = "PSA Backed Double-D Single Hole Gland Calculator"
    sheet["B2"].font = Font(bold=True, size=13, color="2323FF")
    sheet["B3"] = "Model after Abraham Lee, IPPD (github.com/tisimst/vatic)"

    sheet["B6"] = "INPUTS"
    sheet["B6"].font = _HEADING
    sheet["C6"] = "Nominal values in the shaded fields"
    for column, title in (
        ("B", "Parameter"),
        ("C", "Nominal"),
        ("D", "Tolerance (+/-)"),
        ("E", "Units"),
    ):
        cell = sheet[f"{column}7"]
        cell.value = title
        cell.font = _BOLD

    for row, label, nominal, tolerance, units, name in INPUTS:
        sheet[f"B{row}"] = label
        sheet[f"C{row}"] = nominal
        sheet[f"C{row}"].fill = _INPUT_FILL
        sheet[f"D{row}"] = tolerance
        sheet[f"E{row}"] = units
        book.defined_names.add(
            DefinedName(name, attr_text=f"'{SHEET}'!$C${row}")
        )

    sheet["B17"] = "REQUIREMENTS"
    sheet["B17"].font = _HEADING
    for column, title in (
        ("B", "Parameter"),
        ("C", "Lower Limit"),
        ("D", "Upper Limit"),
    ):
        cell = sheet[f"{column}18"]
        cell.value = title
        cell.font = _BOLD

    for offset, (row, _label, _formula, lower, upper) in enumerate(OUTPUTS):
        spec_row = 19 + offset
        sheet[f"B{spec_row}"] = f"={'B'}{row}"
        sheet[f"C{spec_row}"] = lower
        sheet[f"D{spec_row}"] = upper

    sheet["B22"] = "INTERMEDIATE CALCULATIONS (REFERENCE)"
    sheet["B22"].font = _HEADING
    for column, title in (("B", "Parameter"), ("C", "Nominal"), ("D", "Units")):
        cell = sheet[f"{column}23"]
        cell.value = title
        cell.font = _BOLD

    for row, label, formula, units, name in INTERMEDIATES:
        sheet[f"B{row}"] = label
        sheet[f"C{row}"] = formula
        sheet[f"D{row}"] = units
        book.defined_names.add(
            DefinedName(name, attr_text=f"'{SHEET}'!$C${row}")
        )

    sheet["B28"] = "OUTPUTS"
    sheet["B28"].font = _HEADING
    sheet["B29"] = "Parameter"
    sheet["B29"].font = _BOLD
    sheet["C29"] = "Nominal"
    sheet["C29"].font = _BOLD

    for row, label, formula, _lower, _upper in OUTPUTS:
        sheet[f"B{row}"] = label
        sheet[f"C{row}"] = formula

    sheet.column_dimensions["B"].width = 30
    for column in ("C", "D", "E"):
        sheet.column_dimensions[column].width = 16
    for row in sheet.iter_rows(min_row=1, max_row=32, max_col=5):
        for cell in row:
            if cell.column_letter in {"C", "D"}:
                cell.alignment = Alignment(horizontal="right")

    book.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = build()
    print(f"wrote {path}")
